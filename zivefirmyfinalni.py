import json
import logging
import random
import re
import subprocess
import time
import urllib.parse
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from threading import Lock
from typing import Dict, List, Optional
from urllib.parse import urljoin, urlparse
from thefuzz import fuzz
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import PatternFill, Alignment
from playwright.sync_api import sync_playwright
from requests.adapters import HTTPAdapter
from tqdm import tqdm
from urllib3.util.retry import Retry
import json
import logging
import requests
from dataclasses import dataclass
from datetime import datetime
from typing import Optional, List
import os

# Nastavení loggeru
logging.basicConfig(
    filename='ares_scraper.log',
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    encoding='utf-8'
)

logger = logging.getLogger('AresScraper')


@dataclass
class RetrySession:
    """Třída pro vytvoření session s automatickým opakováním požadavků"""
    
    def __init__(self, retries=3, backoff_factor=0.3, 
                status_forcelist=(500, 502, 504), 
                allowed_methods=frozenset(['GET', 'HEAD', 'OPTIONS'])):
        self.session = requests.Session()
        
        # Explicitní vytvoření instance Retry
        retry_strategy: Retry = Retry(
            total=retries,
            read=retries,
            connect=retries,
            backoff_factor=backoff_factor,
            status_forcelist=status_forcelist,
            allowed_methods=allowed_methods
        )
        
        adapter = HTTPAdapter(max_retries=retry_strategy)
        self.session.mount('http://', adapter)
        self.session.mount('https://', adapter)


@dataclass
class AresCompanyData:
    ico: str
    nazev: str
    pravni_forma: Optional[str] = None
    datum_vzniku: Optional[str] = None
    sidlo: Optional[str] = None
    zamestnanci: Optional[str] = None
    sidlo_ares: Optional[str] = None
    datumZaniku: Optional[str] = None
    email: Optional[str] = None
    web: Optional[str] = None
@dataclass
class WebsiteLanguageInfo:
    main_domain: str
    language_versions: Dict[str, str]  # jazyk: URL
    detected_languages: List[str]      # jazyky detekované na stránce
    language_switcher_present: bool
    
class WebsiteLanguageAnalyzer:
    def __init__(self, logger=None):
        self.logger = logger or logging.getLogger(__name__)
        self._setup_logging()
        
        # Mapování TLD na jazyky
        self.TLD_LANGUAGE_MAP = {
            'cz': 'cs',
            'sk': 'sk',
            'de': 'de',
            'at': 'de',
            'ch': 'de',
            'uk': 'en',
            'com': 'en',
            'pl': 'pl',
            'hu': 'hu',
            'fr': 'fr',
            'es': 'es',
            'it': 'it',
            'nl': 'nl'
        }

        # Vylepšené specifické vzory pro detekci jazyků
        self.LANGUAGE_PATTERNS = {
            'cs': [
                r'<html[^>]+lang=["\']cs["\']',
                r'<meta[^>]+content=["\'][^"\']*cs[^"\']*["\']',
                r'(?:^|\s|/)(čeština|czech)(?:\s|$|/)',
                r'hreflang=["\']cs["\']'
            ],
            'en': [
                r'<html[^>]+lang=["\']en["\']',
                r'<meta[^>]+content=["\'][^"\']*en[^"\']*["\']',
                r'(?:^|\s|/)(english|anglicky)(?:\s|$|/)',
                r'hreflang=["\']en["\']'
            ],
            'de': [
                r'<html[^>]+lang=["\']de["\']',
                r'<meta[^>]+content=["\'][^"\']*de[^"\']*["\']',
                r'(?:^|\s|/)(deutsch|german|německy)(?:\s|$|/)',
                r'hreflang=["\']de["\']'
            ],
            'sk': [
                r'<html[^>]+lang=["\']sk["\']',
                r'<meta[^>]+content=["\'][^"\']*sk[^"\']*["\']',
                r'(?:^|\s|/)(slovenčina|slovak|slovensky)(?:\s|$|/)',
                r'hreflang=["\']sk["\']'
            ],
            'pl': [
                r'<html[^>]+lang=["\']pl["\']',
                r'<meta[^>]+content=["\'][^"\']*pl[^"\']*["\']',
                r'(?:^|\s|/)(polski|polish|polsky)(?:\s|$|/)',
                r'hreflang=["\']pl["\']'
            ]
        }

        # Vzory pro jazykové přepínače
        self.LANGUAGE_SWITCHER_PATTERNS = [
            r'class=["\'].*?lang.*?switch.*?["\']',
            r'id=["\'].*?lang.*?switch.*?["\']',
            r'class=["\'].*?language.*?selector.*?["\']',
            r'<select[^>]*?name=["\'].*?lang.*?["\']'
        ]

    def _setup_logging(self):
        """Nastavení loggeru"""
        if not self.logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
            handler.setFormatter(formatter)
            self.logger.addHandler(handler)
            self.logger.setLevel(logging.INFO)

    def analyze_website_languages(self, url: str) -> WebsiteLanguageInfo:
        """Hlavní metoda pro analýzu jazykových verzí webu"""
        self.logger.info(f"Začínám analýzu jazyků pro web: {url}")
        try:
            main_domain = self._get_main_domain(url)
            self.logger.info(f"Získána hlavní doména: {main_domain}")
            
            language_versions = self._check_language_versions(main_domain)
            self.logger.info(f"Nalezené jazykové verze podle TLD: {language_versions}")
            
            with sync_playwright() as p:
                self.logger.info("Spouštím Playwright...")
                browser = p.chromium.launch(headless=True)
                page = browser.new_page()
                
                self.logger.info(f"Načítám stránku: {url}")
                page.goto(url, wait_until='networkidle')
                
                detected_languages = self._detect_page_languages(page)
                self.logger.info(f"Detekované jazyky na stránce: {detected_languages}")
                
                has_language_switcher = self._check_language_switcher(page)
                self.logger.info(f"Přítomnost jazykového přepínače: {has_language_switcher}")
                
                if has_language_switcher:
                    self.logger.info("Hledám další jazykové verze v přepínači...")
                    additional_versions = self._get_language_versions_from_switcher(page)
                    self.logger.info(f"Nalezené dodatečné jazykové verze: {additional_versions}")
                    language_versions.update(additional_versions)
                
                browser.close()
                self.logger.info("Playwright ukončen")
                
            return WebsiteLanguageInfo(
                main_domain=main_domain,
                language_versions=language_versions,
                detected_languages=detected_languages,
                language_switcher_present=has_language_switcher
            )
            
        except Exception as e:
            self.logger.error(f"Chyba při analýze webu {url}: {str(e)}", exc_info=True)
            return WebsiteLanguageInfo(
                main_domain=url,
                language_versions={},
                detected_languages=[],
                language_switcher_present=False
            )

    def _get_main_domain(self, url: str) -> str:
        """Získá hlavní doménu z URL"""
        self.logger.debug(f"Získávám hlavní doménu z URL: {url}")
        try:
            parsed = urlparse(url)
            self.logger.debug(f"Parsované URL: {parsed}")
            domain_parts = parsed.netloc.split('.')
            self.logger.debug(f"Části domény: {domain_parts}")
            
            # Odstranění www. pokud existuje
            if domain_parts[0] == 'www':
                domain_parts = domain_parts[1:]
                
            # Získání hlavní části domény
            main_domain = domain_parts[-2]  # např. pro firma.cz získáme "firma"
            self.logger.debug(f"Získaná hlavní doména: {main_domain}")
            return main_domain
        except Exception as e:
            self.logger.error(f"Chyba při získávání hlavní domény: {str(e)}")
            return ""

    def _check_language_versions(self, main_domain: str) -> Dict[str, str]:
        """Kontrola existence různých TLD variant domény"""
        self.logger.info(f"Kontroluji jazykové verze pro doménu: {main_domain}")
        language_versions = {}
        
        for tld, language in self.TLD_LANGUAGE_MAP.items():
            test_url = f"https://{main_domain}.{tld}"
            self.logger.debug(f"Testuji URL: {test_url}")
            try:
                response = requests.head(test_url, timeout=5, allow_redirects=True)
                self.logger.debug(f"Status code pro {test_url}: {response.status_code}")
                if response.status_code == 200:
                    language_versions[language] = test_url
                    self.logger.info(f"Nalezena jazyková verze: {language} -> {test_url}")
                time.sleep(random.uniform(0.5, 1))
            except Exception as e:
                self.logger.debug(f"Chyba při testování {test_url}: {str(e)}")
                continue
                    
        return language_versions

    def _detect_page_languages(self, page) -> set[str]:
        """Vylepšená detekce jazyků na stránce s přísnějšími kritérii"""
        detected_languages = set()
        content = page.content().lower()
        
        # 1. Kontrola HTML lang atributu (vysoká důvěryhodnost)
        html_lang = page.evaluate('() => document.documentElement.lang')
        if html_lang:
            lang_code = html_lang.split('-')[0].lower()
            if lang_code in self.LANGUAGE_PATTERNS:
                detected_languages.add(lang_code)
                
        # 2. Kontrola hreflang atributů (vysoká důvěryhodnost)
        hreflang_values = page.evaluate('''() => {
            const links = document.querySelectorAll('link[hreflang]');
            return Array.from(links).map(link => link.getAttribute('hreflang'));
        }''')
        for hreflang in hreflang_values:
            lang_code = hreflang.split('-')[0].lower()
            if lang_code in self.LANGUAGE_PATTERNS:
                detected_languages.add(lang_code)

        # 3. Kontrola meta tagů (střední důvěryhodnost)
        meta_langs = page.evaluate('''() => {
            const metas = document.getElementsByTagName('meta');
            const langs = [];
            for (let meta of metas) {
                if (meta.getAttribute('http-equiv') === 'content-language' ||
                    meta.getAttribute('name') === 'language') {
                    langs.push(meta.getAttribute('content'));
                }
            }
            return langs;
        }''')
        
        for meta_lang in meta_langs:
            if meta_lang:
                lang_code = meta_lang.split('-')[0].lower()
                if lang_code in self.LANGUAGE_PATTERNS:
                    detected_languages.add(lang_code)

        # 4. Kontrola jazykového přepínače (vysoká důvěryhodnost)
        if self._has_language_switcher(page):
            lang_links = self._get_language_links(page)
            for lang in lang_links:
                if lang in self.LANGUAGE_PATTERNS:
                    detected_languages.add(lang)

        # 5. Kontrola URL struktury (střední důvěryhodnost)
        current_url = page.url
        url_parts = urlparse(current_url).path.lower().split('/')
        for part in url_parts:
            if part in ['cs', 'en', 'de', 'sk', 'pl']:
                detected_languages.add(part)

        # 6. Dodatečná validace na základě specifických vzorů
        for lang, patterns in self.LANGUAGE_PATTERNS.items():
            pattern_matches = 0
            for pattern in patterns:
                if re.search(pattern, content, re.IGNORECASE):
                    pattern_matches += 1
            # Vyžadujeme alespoň 2 shody vzorů pro přidání jazyka
            if pattern_matches >= 2:
                detected_languages.add(lang)

        return detected_languages

    def _get_language_links(self, page) -> set[str]:
        """Získání jazyků z odkazů v přepínači"""
        detected_langs = set()
        
        # Hledání odkazů s jazykovými kódy
        links = page.evaluate('''() => {
            const languageLinks = document.querySelectorAll('a[href*="/cs/"], a[href*="/en/"], a[href*="/de/"], a[href*="/sk/"], a[href*="/pl/"]');
            return Array.from(languageLinks).map(link => link.href);
        }''')
        
        for link in links:
            for lang in ['cs', 'en', 'de', 'sk', 'pl']:
                if f'/{lang}/' in link.lower():
                    detected_langs.add(lang)
        
        return detected_langs
    
    def _validate_language_content(self, text: str) -> bool:
        """Validace, že text obsahuje skutečné jazykové volby"""
        language_indicators = [
            r'(?:čeština|česky|czech)',
            r'(?:english|anglicky|en)',
            r'(?:deutsch|německy|german)',
            r'(?:slovenčina|slovensky|slovak)',
            r'(?:polski|polsky|polish)',
            r'(?:language|jazyk|sprache|langue|idioma)',
            r'(?:select.*language|zvolte.*jazyk|wählen.*sprache)',
            r'(?:lang|sprache|langue|idioma)'
        ]
        
        matches = 0
        for indicator in language_indicators:
            if re.search(indicator, text, re.IGNORECASE):
                matches += 1
        
        # Vyžadujeme alespoň dvě jazykové volby pro potvrzení
        return matches >= 2
    
    def _has_language_switcher(self, page) -> bool:
        """Kontrola přítomnosti jazykového přepínače s přísnějšími kritérii"""
        content = page.content()
        
        # Kontrola běžných implementací jazykových přepínačů
        for pattern in self.LANGUAGE_SWITCHER_PATTERNS:
            if re.search(pattern, content, re.IGNORECASE):
                # Dodatečná validace - ověření, že přepínač obsahuje skutečné jazykové volby
                surrounding_text = re.findall(r'.{0,50}' + pattern + r'.{0,50}', content)
                if any(self._validate_language_content(text) for text in surrounding_text):
                    return True
        
        # Kontrola přítomnosti typických jazykových odkazů
        language_links = page.evaluate('''() => {
            const links = document.querySelectorAll('a[href*="/cs/"], a[href*="/en/"], a[href*="/de/"]');
            return links.length > 0;
        }''')
        
        return bool(language_links)

    def _check_language_switcher(self, page) -> bool:
        """Kontroluje přítomnost jazykového přepínače"""
        try:
            # Místo přímého použití regexů jako CSS selektorů použijeme bezpečnější selektory
            language_selectors = [
                # Běžné třídy pro jazykový přepínač
                '[class*="lang-switch"]',
                '[class*="language-switch"]',
                '[class*="lang-selector"]',
                '[class*="language-selector"]',
                '[class*="lang-menu"]',
                '[class*="language-menu"]',
                
                # ID pro jazykový přepínač
                '[id*="lang-switch"]',
                '[id*="language-switch"]',
                '[id*="lang-selector"]',
                '[id*="language-selector"]',
                
                # Obecné selektory pro výběr jazyka
                'select[name*="lang"]',
                'select[id*="lang"]',
                'div[class*="lang"] a',
                'ul[class*="lang"] a',
                
                # Specifické elementy pro jazykové verze
                '.language-chooser',
                '.lang-chooser',
                '.language-list',
                '.lang-list',
                
                # Další běžné konvence
                '[aria-label*="language"]',
                '[aria-label*="Languages"]',
                '[data-lang-switcher]',
                '[data-language-switcher]'
            ]
            
            # Spojíme všechny selektory do jednoho
            combined_selector = ', '.join(language_selectors)
            
            # Kontrola přítomnosti jakéhokoliv z těchto elementů
            language_elements = page.query_selector_all(combined_selector)
            
            # Dodatečná kontrola pomocí evaluace JavaScriptu pro složitější případy
            has_language_elements = page.evaluate('''() => {
                // Hledání elementů obsahujících typické texty pro jazykové přepínače
                const languageTexts = ['language', 'jazyk', 'sprache', 'langue', 'idioma'];
                const elements = document.querySelectorAll('a, button, div, span');
                
                for (const element of elements) {
                    const text = element.textContent.toLowerCase();
                    if (languageTexts.some(langText => text.includes(langText))) {
                        return true;
                    }
                }
                
                // Kontrola přítomnosti hreflang atributů
                const hasHreflang = document.querySelector('link[hreflang]') !== null;
                
                return hasHreflang;
            }''')
            
            return bool(language_elements.length > 0 or has_language_elements)
                
        except Exception as e:
            self.logger.error(f"Chyba při kontrole jazykového přepínače: {str(e)}")
            return False

    def _get_language_versions_from_switcher(self, page) -> Dict[str, str]:
        """Získá URL jazykových verzí z přepínače"""
        try:
            # Hledání odkazů na jazykové verze
            language_links = page.evaluate(r'''() => {
                const langLinks = {};
                const selectors = [
                    'a[href*="/de/"]',
                    'a[href*="/en/"]',
                    'a[href*="/cs/"]',
                    'a[href*="/sk/"]',
                    'a[href*="/pl/"]',
                    'a[href*="/hu/"]',
                    'a[href*="/fr/"]',
                    'a[href*="/es/"]',
                    'a[href*="/it/"]'
                ];
                
                document.querySelectorAll(selectors.join(', ')).forEach(link => {
                    const href = link.href;
                    const lang = href.match(/\/([a-z]{2})\//i)?.[1];
                    if (lang) {
                        langLinks[lang.toLowerCase()] = href;
                    }
                });
                
                return langLinks;
            }''')
            
        except Exception as e:
            self.logger.error(f"Chyba při získávání jazykových verzí z přepínače: {str(e)}")
            return {}

class ExportMarketsAnalyzer:
    """Třída pro analýzu exportních trhů z webových stránek"""
    
    def __init__(self, logger=None):
        self.logger = logger or logging.getLogger(__name__)
        
        # Definice vzorů pro evropské země a regiony
        self.EXPORT_PATTERNS = {
            # Střední Evropa
            'Německo': [
                r'německo', r'německý\s*trh', r'německé\s*firmy',
                r'deutschland', r'deutschen?\s*markt', 
                r'germany', r'german\s*market',
                r'do\s*německa', r'v\s*německu',
                r'německým\s*zákazníkům',
                r'německým\s*partnerům',
                r'německy\s*mluvících\s*zemích',
                r'spolupráce\s*s\s*německými'
            ],
            'Rakousko': [
                r'rakousko', r'rakouský\s*trh',
                r'österreich', r'austria',
                r'do\s*rakouska', r'v\s*rakousku',
                r'rakouským\s*zákazníkům',
                r'vídeň', r'wien'
            ],
            'Polsko': [
                r'polsko', r'polský\s*trh',
                r'polska', r'poland', r'polish\s*market',
                r'do\s*polska', r'v\s*polsku',
                r'polským\s*zákazníkům',
                r'varšava', r'warszawa'
            ],
            'Slovensko': [
                r'slovensko', r'slovenský\s*trh',
                r'slovakia', r'slovak\s*market',
                r'do\s*slovenska', r'na\s*slovensku',
                r'slovenským\s*zákazníkům',
                r'bratislava'
            ],
            'Maďarsko': [
                r'maďarsko', r'maďarský\s*trh',
                r'hungary', r'hungarian\s*market',
                r'do\s*maďarska', r'v\s*maďarsku',
                r'budapest'
            ],
            
            # Západní Evropa
            'Francie': [
                r'francie', r'francouzský\s*trh',
                r'france', r'french\s*market',
                r'do\s*francie', r've\s*francii',
                r'francouzským\s*zákazníkům',
                r'paříž', r'paris'
            ],
            'Benelux': [
                r'benelux',
                r'nizozemsko', r'nizozemský\s*trh',
                r'belgie', r'belgický\s*trh',
                r'netherlands', r'belgium', r'dutch\s*market',
                r'do\s*nizozemska', r'do\s*belgie',
                r'holandsko', r'do\s*holandska',
                r'amsterdam', r'brusel', r'rotterdam'
            ],
            'Velká Británie': [
                r'británie', r'britský\s*trh',
                r'anglie', r'uk\s*market',
                r'united\s*kingdom', r'great\s*britain',
                r'do\s*británie', r'v\s*británii',
                r'do\s*anglie', r'v\s*anglii',
                r'londýn', r'london'
            ],
            
            # Severní Evropa
            'Skandinávie': [
                r'skandinávie', r'skandinávský\s*trh',
                r'švédsko', r'švédský\s*trh',
                r'norsko', r'norský\s*trh',
                r'dánsko', r'dánský\s*trh',
                r'finsko', r'finský\s*trh',
                r'scandinavia', r'scandinavian\s*market',
                r'sweden', r'norway', r'denmark', r'finland',
                r'do\s*švédska', r've\s*švédsku',
                r'do\s*norska', r'v\s*norsku',
                r'do\s*dánska', r'v\s*dánsku',
                r'do\s*finska', r've\s*finsku',
                r'stockholm', r'oslo', r'kodaň', r'helsinki'
            ],
            
            # Jižní Evropa
            'Itálie': [
                r'itálie', r'italský\s*trh',
                r'italy', r'italian\s*market',
                r'do\s*itálie', r'v\s*itálii',
                r'italským\s*zákazníkům',
                r'řím', r'milano', r'rome'
            ],
            'Španělsko': [
                r'španělsko', r'španělský\s*trh',
                r'spain', r'spanish\s*market',
                r'do\s*španělska', r've\s*španělsku',
                r'madrid', r'barcelona'
            ]
        }
        
        
       

    def analyze_export_markets(self, page) -> Dict[str, any]:
        try:
            content = page.content().lower()
            text = page.inner_text('body').lower()
            
            for pattern in self.NEGATIVE_PATTERNS:
                if re.search(pattern, text, re.IGNORECASE):
                    self.logger.info(f"Nalezen negativní vzor: {pattern}")
                    return {
                        'export_markets': [],
                        'export_percentage': None,
                        'has_export_activity': False,
                        'export_evidence': [],
                        'main_markets': [],
                        'is_transport_company': True
                    }
            
            result = {
                'export_markets': [],
                'export_percentage': None,
                'has_export_activity': False,
                'export_evidence': [],
                'main_markets': [],
                'is_transport_company': False
            }
            # Kontrola exportní činnosti
            for pattern in self.EXPORT_ACTIVITY_PATTERNS:
                matches = re.finditer(pattern, text, re.IGNORECASE)
                for match in matches:
                    context = self._get_context(text, match.start(), 100)
                    if context:
                        result['export_evidence'].append(context)
                        result['has_export_activity'] = True
            
            # Detekce zemí a regionů
            market_mentions = {}
            for market, patterns in self.EXPORT_PATTERNS.items():
                mentions = []
                for pattern in patterns:
                    matches = re.finditer(pattern, text, re.IGNORECASE)
                    for match in matches:
                        context = self._get_context(text, match.start(), 100)
                        if context and self._is_valid_export_context(context):
                            mentions.append(context)
                
                if mentions:
                    market_mentions[market] = len(mentions)
                    result['export_markets'].append({
                        'market': market,
                        'mentions': len(mentions),
                        'evidence': mentions[:3]  # Ukládáme max 3 příklady zmínek
                    })
            
            # Identifikace hlavních trhů (podle počtu zmínek)
            if market_mentions:
                sorted_markets = sorted(market_mentions.items(), key=lambda x: x[1], reverse=True)
                result['main_markets'] = [market for market, _ in sorted_markets[:3]]
            
            return result
            
        except Exception as e:
            self.logger.error(f"Chyba při analýze exportních trhů: {str(e)}")
            return None

    def _get_context(self, text: str, position: int, context_size: int) -> str:
        """Získá okolní kontext pro danou pozici v textu"""
        start = max(0, position - context_size)
        end = min(len(text), position + context_size)
        return text[start:end].strip()

    def _is_valid_export_context(self, context: str) -> bool:
        """Ověří, zda je kontext skutečně spojen s exportem"""
        positive_indicators = [
            'export', 'dodáv', 'vývoz', 'působ', 'zastoup',
            'pobočk', 'zákazník', 'trh', 'market', 'deliver'
        ]
        negative_indicators = [
            'import', 'dovoz', 'nakupujeme', 'odebíráme'
        ]
        
        # Kontrola pozitivních indikátorů
        has_positive = any(indicator in context.lower() for indicator in positive_indicators)
        # Kontrola negativních indikátorů
        has_negative = any(indicator in context.lower() for indicator in negative_indicators)
        
        return has_positive and not has_negative

class ContactedCompaniesFilter:
    def __init__(self, contacted_companies_file):
        self.contacted_companies = self._load_contacted_companies(contacted_companies_file)
        self.logger = logging.getLogger('ContactedCompaniesFilter')
        # Přidáme log o načtených firmách
        self.logger.info(f"Načteno {len(self.contacted_companies)} kontaktovaných firem:")
        for company in self.contacted_companies:
            self.logger.info(f"Načtená firma (normalizovaná): {company}")
        
    def _load_contacted_companies(self, file_path):
        """Načte seznam již kontaktovaných firem z Excel souboru"""
        try:
            df = pd.read_excel(file_path)
            # Předpokládáme, že názvy firem jsou v prvním sloupci
            company_names = set()
            
            for name in df.iloc[:, 0]:  # První sloupec
                if isinstance(name, str):
                    # Normalizace názvu
                    normalized_name = self._normalize_company_name(name)
                    company_names.add(normalized_name)
                    
            return company_names
            
        except Exception as e:
            self.logger.error(f"Chyba při načítání seznamu kontaktovaných firem: {e}")
            return set()
            
    def _normalize_company_name(self, name):
        """Normalizuje název firmy pro konzistentní porovnávání"""
        if not isinstance(name, str):
            return ""
            
        # Odstranění právních forem
        replacements = {
            'spol. s r.o.': '',
            'spol.s r.o.': '',
            'spol s.r.o.': '',
            'spol.s.r.o.': '',
            's.r.o.': '',
            'a.s.': '',
            'v.o.s.': '',
            'k.s.': '',
            'z.s.': '',
            'o.p.s.': '',
            'družstvo': ''
        }
        
        normalized = name.lower()
        for old, new in replacements.items():
            normalized = normalized.replace(old, new)
            
        # Odstranění diakritiky
        normalized = normalized.translate(str.maketrans('áčďéěíňóřšťúůýž', 'acdeeinorstuuyz'))
        
        # Odstranění speciálních znaků a nadbytečných mezer
        normalized = ''.join(c for c in normalized if c.isalnum() or c.isspace())
        normalized = ' '.join(normalized.split())
        
        return normalized
        
    def is_company_contacted(self, company_name, similarity_threshold=85):
        """Zkontroluje, zda byla firma již kontaktována"""
        normalized_name = self._normalize_company_name(company_name)
        self.logger.info(f"Kontroluji firmu: {company_name}")
        self.logger.info(f"Normalizovaný název: {normalized_name}")
        
        # Nejdřív zkusíme přesnou shodu
        if normalized_name in self.contacted_companies:
            self.logger.info(f"Nalezena přesná shoda pro: {company_name}")
            return True
            
        # Pokud není přesná shoda, zkusíme fuzzy matching
        for contacted in self.contacted_companies:
            similarity = fuzz.ratio(normalized_name, contacted)
            self.logger.info(f"Porovnávám '{normalized_name}' s '{contacted}' - podobnost: {similarity}%")
            if similarity >= similarity_threshold:
                self.logger.info(f"Nalezena podobnost {similarity}% pro: {company_name}")
                return True
                
        self.logger.info(f"Firma '{company_name}' nebyla nalezena v seznamu kontaktovaných")
        return False
              
class WebAnalyzer:
    """Vylepšená třída pro analýzu webových stránek"""
    
    def __init__(self, logger, max_retries=3, timeout=30):
        self.logger = logger
        self.max_retries = max_retries
        self.timeout = timeout
        self._setup_playwright()
        self.language_analyzer = WebsiteLanguageAnalyzer(logger)
        self.export_analyzer = ExportMarketsAnalyzer(logger)

    def extract_deep_emails(self, page) -> set:
        """Rozšířená extrakce emailů z webu včetně skrytých a obfuskovaných emailů"""
        found_emails = set()
        
        try:
            # 1. Získání všech textů ze stránky včetně skrytých elementů
            all_text = page.evaluate('''() => {
                function getAllText(element) {
                    let text = '';
                    // Zahrnout komentáře, které mohou obsahovat skryté emaily
                    for (let node of element.childNodes) {
                        if (node.nodeType === Node.COMMENT_NODE) {
                            text += node.textContent + ' ';
                        }
                    }
                    // Získat computed style elementu
                    let style = window.getComputedStyle(element);
                    // Kontrola viditelnosti
                    if (style.display !== 'none' && style.visibility !== 'hidden') {
                        text += element.innerText + ' ';
                    }
                    return text;
                }
                return Array.from(document.getElementsByTagName('*'))
                    .map(getAllText)
                    .join(' ');
            }''')

            # 2. Skenování různých formátů emailů
            email_patterns = [
                # Standardní emaily
                r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}',
                # Emaily s textem "at" a "dot"
                r'[a-zA-Z0-9._%+-]+\s+(?:at|@|AT)\s+[a-zA-Z0-9.-]+\s+(?:dot|\.|\(dot\))\s+[a-zA-Z]{2,}',
                # Emaily v HTML entitách
                r'[a-zA-Z0-9._%+-]+&#(?:64|0*64);[a-zA-Z0-9.-]+&#(?:46|0*46);[a-zA-Z]{2,}',
                # Emaily s Unicode znaky
                r'[a-zA-Z0-9._%+-]+(?:[@＠]{1})[a-zA-Z0-9.-]+[.．。][a-zA-Z]{2,}'
            ]

            for pattern in email_patterns:
                matches = re.finditer(pattern, all_text, re.IGNORECASE)
                for match in matches:
                    email = match.group()
                    # Čištění a normalizace emailu
                    email = self._normalize_email(email)
                    if self._is_valid_email(email):
                        found_emails.add(email)

            # 3. Hledání v JavaScript souborech
            js_content = page.evaluate('''() => {
                return Array.from(document.getElementsByTagName('script'))
                    .map(script => script.textContent)
                    .join(' ');
            }''')
            
            for pattern in email_patterns:
                matches = re.finditer(pattern, js_content, re.IGNORECASE)
                for match in matches:
                    email = self._normalize_email(match.group())
                    if self._is_valid_email(email):
                        found_emails.add(email)

            # 4. Hledání v data atributech
            data_attrs = page.evaluate('''() => {
                return Array.from(document.getElementsByTagName('*'))
                    .map(el => Array.from(el.attributes)
                        .filter(attr => attr.name.startsWith('data-'))
                        .map(attr => attr.value)
                        .join(' ')
                    )
                    .join(' ');
            }''')

            for pattern in email_patterns:
                matches = re.finditer(pattern, data_attrs, re.IGNORECASE)
                for match in matches:
                    email = self._normalize_email(match.group())
                    if self._is_valid_email(email):
                        found_emails.add(email)

            # 5. Kontrola odkazů na kontaktní stránky
            contact_links = page.evaluate('''() => {
                return Array.from(document.querySelectorAll('a[href*="kontakt"], a[href*="contact"]'))
                    .map(a => a.href);
            }''')

            for link in contact_links:
                try:
                    page.goto(link, wait_until='domcontentloaded', timeout=10000)
                    contact_text = page.content()
                    for pattern in email_patterns:
                        matches = re.finditer(pattern, contact_text, re.IGNORECASE)
                        for match in matches:
                            email = self._normalize_email(match.group())
                            if self._is_valid_email(email):
                                found_emails.add(email)
                except Exception as e:
                    self.logger.debug(f"Nelze načíst kontaktní stránku {link}: {str(e)}")

            return found_emails

        except Exception as e:
            self.logger.error(f"Chyba při hledání emailů: {str(e)}")
            return found_emails

    def _normalize_email(self, email: str) -> str:
        """Normalizuje nalezený email"""
        # Odstranění bílých znaků
        email = email.strip().lower()
        
        # Nahrazení textových reprezentací znaků
        replacements = {
            ' at ': '@',
            ' dot ': '.',
            '(at)': '@',
            '[at]': '@',
            '(dot)': '.',
            '[dot]': '.',
            '＠': '@',
            '．': '.',
            '。': '.'
        }
        
        for old, new in replacements.items():
            email = email.replace(old, new)
        
        # Odstranění HTML entit
        email = re.sub(r'&#(?:64|0*64);', '@', email)
        email = re.sub(r'&#(?:46|0*46);', '.', email)
        
        # Odstranění dalších nežádoucích znaků
        email = re.sub(r'[^\w\.-@]', '', email)
        
        return email

    def _is_valid_email(self, email: str) -> bool:
        """Rozšířená validace emailové adresy"""
        if not email or '@' not in email:
            return False
            
        # Základní pattern pro validaci emailu
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        
        # Kontrola délky
        if len(email) < 5 or len(email) > 254:
            return False
            
        # Kontrola formátu
        if not re.match(pattern, email):
            return False
            
        # Kontrola domény
        domain = email.split('@')[1]
        if len(domain) < 3 or '.' not in domain:
            return False
            
        # Blacklist obecných/testovacích emailů
        blacklist = {
            'example.com', 'test.com', 'domain.com', 
            'email.com', 'website.com', 'yoursite.com'
        }
        if domain in blacklist:
            return False
            
        return True
        
    
    def extract_contacts(self, page) -> dict:
        """Extrahuje telefonní čísla a emaily z webové stránky"""
        try:
            # Původní extrakce telefonů zůstává stejná
            found_numbers = set()
            text = page.inner_text('body')
            page_content = page.content()
                
            phone_patterns = [
                r'(?:tel\.?|telefon|mobil)[:\s.]*(?:\+|00)\s*(?:420|421)\s*\d{3}\s*\d{3}\s*\d{3}',
                r'(?:tel\.?|telefon|mobil)[:\s.]*\d{3}\s*\d{3}\s*\d{3}',
                r'\+(?:420|421)\s*\d{3}\s*\d{3}\s*\d{3}',
                r'00(?:420|421)\s*\d{3}\s*\d{3}\s*\d{3}',
                r'(?<!\d)\d{3}[-\s]*\d{3}[-\s]*\d{3}(?!\d)',
                r'tel(?:efon)?\.?[:. ]+([+\d]\d{2,}[\d\s-]{8,})',
                r'mob(?:il)?\.?[:. ]+([+\d]\d{2,}[\d\s-]{8,})'
            ]
                
            for pattern in phone_patterns:
                matches = re.finditer(pattern, text, re.IGNORECASE | re.MULTILINE)
                for match in matches:
                    number = match.group().strip()
                    number = re.sub(r'^(?:tel\.?|telefon|mobil)[:\s.]+', '', number, flags=re.IGNORECASE)
                    cleaned = re.sub(r'[^\d+]', '', number)
                        
                    if cleaned.startswith('00420'):
                        cleaned = '+420' + cleaned[5:]
                    elif cleaned.startswith('00421'):
                        cleaned = '+421' + cleaned[5:]
                    elif len(cleaned) == 9 and cleaned.isdigit():
                        cleaned = '+420' + cleaned
                        
                    if len(cleaned) >= 13 and len(cleaned) <= 14:
                        if (cleaned.startswith('+420') or cleaned.startswith('+421')):
                            rest_of_number = cleaned[4:]
                            if len(rest_of_number) == 9 and rest_of_number.isdigit():
                                found_numbers.add(cleaned)

            # Nová vylepšená extrakce emailů
            found_emails = self.extract_deep_emails(page)

            return {
                'phone_numbers': sorted(list(found_numbers)),
                'emails': sorted(list(found_emails))
            }

        except Exception as e:
            self.logger.error(f"Chyba při extrakci kontaktů: {str(e)}")
            return {'phone_numbers': [], 'emails': []}

    def _is_valid_email(self, email: str) -> bool:
        """Validace emailové adresy"""
        if not email or '@' not in email:
            return False
            
        # Základní pattern pro validaci emailu
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        
        # Kontrola délky
        if len(email) < 5 or len(email) > 254:
            return False
            
        # Kontrola formátu
        if not re.match(pattern, email):
            return False
            
        # Kontrola domény
        domain = email.split('@')[1]
        if len(domain) < 3 or '.' not in domain:
            return False
            
        return True

    def _setup_playwright(self):
        try:
            subprocess.run(['playwright', 'install'], check=True)
            self.logger.info("Playwright successfully initialized")
        except Exception as e:
            self.logger.error(f"Error initializing Playwright: {str(e)}")

    def validate_language_versions(self, main_url: str, company_name: str, language_versions: dict) -> dict:
        """Validuje, zda jazykové verze patří stejné firmě"""
        validated_versions = {}
        
        try:
            # Získáme hlavní charakteristiky původního webu
            main_characteristics = self._get_website_characteristics(main_url, company_name)
            if not main_characteristics:
                return {}
                
            for language, url in language_versions.items():
                score = 0
                max_score = 5  # Maximální skóre pro shodu
                
                # Získáme charakteristiky jazykové verze
                lang_characteristics = self._get_website_characteristics(url, company_name)
                if not lang_characteristics:
                    continue
                    
                # 1. Kontrola stejné domény nebo subdomény
                if self._is_same_domain(main_url, url):
                    score += 1
                    
                # 2. Kontrola podobnosti layoutu (počet a struktura elementů)
                if self._compare_layout_similarity(
                    main_characteristics['layout'], 
                    lang_characteristics['layout']
                ) > 0.7:  # 70% podobnost
                    score += 1
                    
                # 3. Kontrola přítomnosti loga nebo názvu firmy
                if self._compare_company_presence(
                    main_characteristics['company_elements'],
                    lang_characteristics['company_elements'],
                    company_name
                ):
                    score += 1
                    
                # 4. Kontrola kontaktních údajů
                if self._compare_contact_info(
                    main_characteristics['contacts'],
                    lang_characteristics['contacts']
                ):
                    score += 1
                    
                # 5. Kontrola meta tagů a hlavičky
                if self._compare_meta_info(
                    main_characteristics['meta_info'],
                    lang_characteristics['meta_info']
                ):
                    score += 1
                    
                # Přidáme URL do validovaných verzí, pokud dosáhla alespoň 60% shody
                confidence = score / max_score
                if confidence >= 0.6:
                    validated_versions[language] = {
                        'url': url,
                        'confidence': confidence
                    }
                    self.logger.info(f"Jazyková verze {language} validována s důvěrou {confidence:.2%}")
                else:
                    self.logger.warning(
                        f"Jazyková verze {language} ({url}) zamítnuta - "
                        f"nedostatečná shoda ({confidence:.2%})"
                    )
                    
            return validated_versions
            
        except Exception as e:
            self.logger.error(f"Chyba při validaci jazykových verzí: {str(e)}")
            return {}
        
    def _get_website_characteristics(self, url: str, company_name: str) -> dict:
        """Získá charakteristické znaky webu pro porovnání"""
        try:
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True)
                page = browser.new_page()
                page.goto(url, wait_until='domcontentloaded', timeout=15000)
                
                characteristics = {
                    # Struktura layoutu (počty různých typů elementů)
                    'layout': page.evaluate('''() => {
                        const elements = document.getElementsByTagName('*');
                        const counts = {};
                        for (let el of elements) {
                            counts[el.tagName] = (counts[el.tagName] || 0) + 1;
                        }
                        return counts;
                    }'''),
                    
                    # Elementy obsahující název firmy
                    'company_elements': page.evaluate('''(company) => {
                        return Array.from(document.getElementsByTagName('*'))
                            .filter(el => el.textContent.includes(company))
                            .map(el => ({
                                tag: el.tagName,
                                class: el.className,
                                id: el.id,
                                role: el.getAttribute('role')
                            }));
                    }''', company_name),
                    
                    # Kontaktní informace
                    'contacts': {
                        'phones': self._extract_phones(page),
                        'emails': self._extract_emails(page)
                    },
                    
                    # Meta informace
                    'meta_info': page.evaluate('''() => {
                        return {
                            title: document.title,
                            meta: Array.from(document.getElementsByTagName('meta'))
                                .map(m => ({
                                    name: m.getAttribute('name'),
                                    content: m.getAttribute('content')
                                })),
                            links: Array.from(document.getElementsByTagName('link'))
                                .map(l => ({
                                    rel: l.getAttribute('rel'),
                                    href: l.getAttribute('href')
                                }))
                        };
                    }''')
                }
                
                browser.close()
                return characteristics
                
        except Exception as e:
            self.logger.error(f"Chyba při získávání charakteristik webu {url}: {str(e)}")
            return None
        
    def _is_same_domain(self, url1: str, url2: str) -> bool:
        """Kontroluje, zda URL patří ke stejné doméně nebo subdoméně"""
        try:
            domain1 = urlparse(url1).netloc.split('.')[-2:]
            domain2 = urlparse(url2).netloc.split('.')[-2:]
            return domain1 == domain2
        except:
            return False
        
    def _compare_layout_similarity(self, layout1: dict, layout2: dict) -> float:
        """Porovná podobnost layoutů dvou stránek"""
        try:
            # Získáme všechny unikátní tagy
            all_tags = set(layout1.keys()) | set(layout2.keys())
            
            # Spočítáme podobnost pro každý tag
            similarities = []
            for tag in all_tags:
                count1 = layout1.get(tag, 0)
                count2 = layout2.get(tag, 0)
                if count1 == 0 and count2 == 0:
                    continue
                max_count = max(count1, count2)
                similarity = 1 - abs(count1 - count2) / max_count
                similarities.append(similarity)
                
            return sum(similarities) / len(similarities) if similarities else 0
            
        except Exception as e:
            self.logger.error(f"Chyba při porovnávání layoutů: {str(e)}")
            return 0
    
    def _compare_company_presence(self, elements1: list, elements2: list, company_name: str) -> bool:
        """Porovná výskyt názvu firmy na obou stránkách"""
        try:
            # Kontrola, zda se název firmy vyskytuje v podobných kontextech
            important_elements1 = [e for e in elements1 if e['tag'] in ['H1', 'H2', 'TITLE']]
            important_elements2 = [e for e in elements2 if e['tag'] in ['H1', 'H2', 'TITLE']]
            
            return len(important_elements1) > 0 and len(important_elements2) > 0
            
        except Exception as e:
            self.logger.error(f"Chyba při porovnávání výskytu názvu firmy: {str(e)}")
            return False
        
    def _compare_contact_info(self, contacts1: dict, contacts2: dict) -> bool:
        """Porovná kontaktní informace mezi stránkami"""
        try:
            # Kontrola, zda se vyskytují stejné kontaktní údaje
            phones1 = set(contacts1['phones'])
            phones2 = set(contacts2['phones'])
            emails1 = set(contacts1['emails'])
            emails2 = set(contacts2['emails'])
            
            # Stačí, když se shoduje alespoň jeden kontakt
            return bool(phones1 & phones2 or emails1 & emails2)
            
        except Exception as e:
            self.logger.error(f"Chyba při porovnávání kontaktních informací: {str(e)}")
            return False
        
    def _compare_meta_info(self, meta1: dict, meta2: dict) -> bool:
        """Porovná meta informace mezi stránkami"""
        try:
            # Kontrola podobnosti meta tagů
            important_meta1 = {m['name']: m['content'] for m in meta1['meta'] 
                            if m['name'] in ['author', 'copyright', 'robots']}
            important_meta2 = {m['name']: m['content'] for m in meta2['meta']
                            if m['name'] in ['author', 'copyright', 'robots']}
            
            return bool(set(important_meta1.items()) & set(important_meta2.items()))
            
        except Exception as e:
            self.logger.error(f"Chyba při porovnávání meta informací: {str(e)}")
            return False
        
    def analyze_website(self, url: str) -> dict:
        """Analyzuje web firmy s vylepšeným zpracováním chyb"""
        result = {
            'languages': [],
            'language_versions': {},
            'phone_numbers': [],
            'emails': [],
            'export_data': {}
        }
        
        for attempt in range(self.max_retries):
            try:
                # Nejdřív analyzujeme jazyky
                self.logger.info("Spouštím analýzu jazyků...")
                language_info = self.language_analyzer.analyze_website_languages(url)
                self.logger.info(f"Výsledek analýzy jazyků: {language_info}")
                
                result['languages'] = language_info.detected_languages
                result['language_versions'] = language_info.language_versions
                
                with sync_playwright() as p:
                    browser = p.chromium.launch(headless=True)
                    context = browser.new_context(
                        viewport={'width': 1920, 'height': 1080},
                        user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                    )
                    page = context.new_page()
                    page.set_default_timeout(self.timeout * 1000)
                    
                    self.logger.info(f"Načítám stránku: {url}")
                    response = page.goto(url)
                    if response.status >= 400:
                        raise Exception(f"HTTP error {response.status}")
                        
                    page.wait_for_load_state("networkidle")
                    
                    # Extrakce kontaktů
                    contacts = self.extract_contacts(page)
                    result['phone_numbers'] = contacts['phone_numbers']
                    result['emails'] = contacts['emails']
                    
                    # Analýza exportních trhů
                    self.logger.info("Spouštím analýzu exportních trhů...")
                    export_data = self.export_analyzer.analyze_export_markets(page)
                    if export_data:
                        result['export_data'] = export_data
                        self.logger.info(f"Nalezeny exportní trhy: {export_data['main_markets']}")
                    
                    browser.close()
                    return result
                        
            except Exception as e:
                self.logger.error(f"Pokus {attempt + 1}/{self.max_retries} selhal pro {url}: {str(e)}")
                if attempt < self.max_retries - 1:
                    time.sleep(random.uniform(2, 5))
                        
        return result
    
    def _find_language_version_url(self, page, lang: str) -> str:
        """Najde URL pro konkrétní jazykovou verzi"""
        # Hledání přímých odkazů na jazykovou verzi
        link = page.evaluate(f'''() => {{
            const link = document.querySelector('a[href*="/{lang}/"], a[hreflang="{lang}"]');
            return link ? link.href : null;
        }}''')
        
        if link:
            return link
            
        # Kontrola aktuální URL pro zjištění, zda již nejsme na dané jazykové verzi
        current_url = page.url
        if f'/{lang}/' in current_url.lower():
            return current_url
            
        return None
    
    def analyze_website_languages(self, url: str) -> dict:
        result = super().analyze_website_languages(url)
        if result['language_versions']:
            # Získáme název firmy z URL nebo z obsahu stránky
            company_name = self._extract_company_name(url)
            # Validujeme jazykové verze
            validated_versions = self.validate_language_versions(
                url, 
                company_name, 
                result['language_versions']
            )
            result['language_versions'] = validated_versions
        return result

class DataExporter:
    """Třída pro export dat do různých formátů"""
    
    def __init__(self, logger):
        self.logger = logger

    def export_to_excel(self, companies: List[dict], filename: str):
        try:
            excel_data = []
            for company in companies:
                # Sloučení a odstranění duplicit emailů
                emails_zivefirmy = set(company.get('emaily', []))
                emails_web = set(company.get('emails', []))
                all_emails = sorted(emails_zivefirmy.union(emails_web))
                
                # Formátování telefonních čísel
                phones_web = sorted(set(company.get('phone_numbers', [])))
                phones_web_formatted = [format_phone_number(phone) for phone in phones_web]
                phones_zivefirmy = sorted(set(company.get('kontaktni_osoby', [])))
                
                # Zpracování exportních dat
                export_data = company.get('export_data', {})
                export_markets = ', '.join(export_data.get('main_markets', []))
                export_percentage = export_data.get('export_percentage', '')
                export_evidence = '; '.join(export_data.get('export_evidence', [])[:3])

                insolvence_info = ''
                if company.get('insolvence'):
                    for ins in company['insolvence']:
                        insolvence_info += (
                            f"Sp. zn.: {ins['spisovaZnacka']}\n"
                            f"Stav: {ins['stavRizeni']}\n"
                            f"Zahájeno: {ins['datumZahajeni']}\n"
                            f"Detail: {ins['urlDetailRizeni']}\n\n"
                        )

                row = {
                    'Název': company.get('nazev', ''),
                    'Insolvence': insolvence_info if insolvence_info else 'Ne',
                    'Právní forma': company.get('pravni_forma', ''),
                    'Počet zaměstnanců': company.get('zamestnanci', ''),
                    'Popis': company.get('popis', ''),
                    'PSČ': company.get('psc', ''),
                    'Adresa': company.get('adresa', ''),
                    'Provozní doba': company.get('provozni_doba', ''),
                    'Telefony zivefirmy': ', '.join(phones_zivefirmy),
                    'Zbylý z webu': ', '.join(phones_web_formatted),
                    'E-maily': ', '.join(all_emails),
                    'Web': ', '.join(company.get('web', [])) if isinstance(company.get('web'), list) else company.get('web', ''),
                    'Jazyky webu': ', '.join(company.get('languages', [])),
                    'Jazykové verze': ', '.join([f"{lang}: {url}" for lang, url in company.get('language_versions', {}).items()]),
                    'Datum vzniku': company.get('datum_vzniku', ''),
                    'IČO': company.get('ic', ''),
                    'Exportní trhy': export_markets,
                    'URL': company.get('url', '')
                }
                excel_data.append(row)

            # Export do Excelu
            df = pd.DataFrame(excel_data)
            
            writer = pd.ExcelWriter(filename, engine='openpyxl')
            df.to_excel(writer, index=False, sheet_name='Firmy')
            
            # Získání pracovního listu
            worksheet = writer.sheets['Firmy']
            
            # Formátování
            for idx, col in enumerate(df.columns):
                # Nastavení šířky sloupce
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(str(col))
                )
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[chr(65 + idx)].width = adjusted_width
                
                # Nastavení barvy hlavičky
                header_cell = worksheet.cell(row=1, column=idx + 1)
                header_cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
                header_cell.alignment = Alignment(horizontal='center')
            
            # Přidání filtru
            worksheet.auto_filter.ref = worksheet.dimensions
            
            # Zmrazení horního řádku
            worksheet.freeze_panes = 'A2'
            
            # Přidání druhého listu se souhrnnými statistikami
            summary_data = {
                'Celkem firem': len(excel_data),
                'Firmy s exportem': len([c for c in excel_data if c['Exportní trhy']]),
                'Nejčastější exportní trhy': self._get_top_markets(excel_data),
                'Firmy s webem': len([c for c in excel_data if c['Web']]),
            }
            
            # Vytvoření listu se souhrnem
            summary_df = pd.DataFrame(list(summary_data.items()), columns=['Metrika', 'Hodnota'])
            summary_df.to_excel(writer, sheet_name='Souhrn', index=False)
            
            # Formátování souhrnného listu
            summary_sheet = writer.sheets['Souhrn']
            for col in ['A', 'B']:
                summary_sheet.column_dimensions[col].width = 30
                header_cell = summary_sheet.cell(row=1, column=ord(col) - ord('A') + 1)
                header_cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
                header_cell.alignment = Alignment(horizontal='center')
                
            writer.close()
            self.logger.info(f"Data úspěšně exportována do souboru: {filename}")
            return True
            
        except Exception as e:
            self.logger.error(f"Chyba při exportu do Excelu: {str(e)}")
            raise
        
    def _get_top_markets(self, data: List[dict], limit: int = 5) -> str:
        """Získá nejčastější exportní trhy"""
        markets = []
        for company in data:
            if company['Exportní trhy']:
                markets.extend(company['Exportní trhy'].split(', '))
        
        if not markets:
            return "Žádné"
            
        market_counts = Counter(markets)
        top_markets = market_counts.most_common(limit)
        return '; '.join([f"{market}: {count}x" for market, count in top_markets])


    def _get_top_technologies(self, data: List[dict], limit: int = 5) -> str:
        """Získá nejčastější technologie"""
        techs = []
        for company in data:
            if company['Technologie']:
                techs.extend(company['Technologie'].split(', '))
        
        if not techs:
            return "Žádné"
            
        tech_counts = Counter(techs)
        top_techs = tech_counts.most_common(limit)
        return '; '.join([f"{tech}: {count}x" for tech, count in top_techs])

    def _get_top_certifications(self, data: List[dict], limit: int = 5) -> str:
        """Získá nejčastější certifikace"""
        certs = []
        for company in data:
            if company['Certifikace']:
                certs.extend(company['Certifikace'].split(', '))
        
        if not certs:
            return "Žádné"
            
        cert_counts = Counter(certs)
        top_certs = cert_counts.most_common(limit)
        return '; '.join([f"{cert}: {count}x" for cert, count in top_certs])

    def export_to_json(self, data: List[dict], filename: str):
        """Export dat do JSON souboru"""
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            self.logger.info(f"Data exportována do {filename}")
            
        except Exception as e:
            self.logger.error(f"Chyba při exportu do JSONu: {str(e)}")
            raise


@dataclass
class AresCompanyData:
    """Datová třída pro ukládání informací o firmě z ARESu"""
    ico: str
    nazev: str
    pravni_forma: Optional[str] = None
    datum_vzniku: Optional[str] = None
    sidlo: Optional[str] = None
    zamestnanci: Optional[str] = None
    sidlo_ares: Optional[str] = None
    datumZaniku: Optional[str] = None

class AresService:
    def __init__(self, logger: logging.Logger):
        self.logger = logger
        self.session = requests.Session()
        retry_strategy = requests.adapters.Retry(
            total=3,
            backoff_factor=1,
            status_forcelist=[429, 500, 502, 503, 504],
        )
        adapter = requests.adapters.HTTPAdapter(max_retries=retry_strategy)
        self.session.mount("https://", adapter)
        self.session.mount("http://", adapter)
        
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'cs,en-US;q=0.9,en;q=0.8'
        })

    def get_company_details(self, ico: str) -> Optional[AresCompanyData]:
        try:
            # 1. Získáme základní data
            basic_url = f"https://ares.gov.cz/ekonomicke-subjekty-v-be/rest/ekonomicke-subjekty/{ico}"
            self.logger.debug(f"Volám základní ARES API: {basic_url}")
            basic_response = self.session.get(basic_url, timeout=30)
            
            if basic_response.status_code == 404:
                self.logger.warning(f"IČO {ico} nebylo nalezeno v ARES")
                return None
                
            basic_response.raise_for_status()
            basic_data = basic_response.json()

            # 2. Získáme data z RES pro počet zaměstnanců
            employees = None
            try:
                res_url = f"https://ares.gov.cz/ekonomicke-subjekty-v-be/rest/ekonomicke-subjekty-res/{ico}"
                self.logger.debug(f"Volám ARES RES API: {res_url}")
                res_response = self.session.get(res_url, timeout=30)
                if res_response.status_code == 200:
                    res_data = res_response.json()
                    
                    if 'zaznamy' in res_data:
                        for zaznam in res_data['zaznamy']:
                            if zaznam.get('primarniZaznam') and 'statistickeUdaje' in zaznam:
                                stat_udaje = zaznam['statistickeUdaje']
                                if 'kategoriePoctuPracovniku' in stat_udaje:
                                    kat_zam = stat_udaje['kategoriePoctuPracovniku']
                                    employees = self._get_employee_category_name(kat_zam)
                                    self.logger.debug(f"Nalezen počet zaměstnanců v RES: {employees}")
                                    break
            except Exception as e:
                self.logger.error(f"Chyba při získávání dat o zaměstnancích z RES: {str(e)}")


            # Formátování sídla
            sidlo_ares = None
            if isinstance(basic_data.get('sidlo'), dict):
                sidlo_parts = []
                address_data = basic_data['sidlo']
                
                if street := address_data.get('nazevUlice'):
                    sidlo_parts.append(street)
                    if num := address_data.get('cisloDomovni'):
                        sidlo_parts[-1] += f" {num}"
                        if orient := address_data.get('cisloOrientacni'):
                            sidlo_parts[-1] += f"/{orient}"
                elif num := address_data.get('cisloDomovni'):
                    sidlo_parts.append(f"č.p. {num}")
                
                if city_part := address_data.get('nazevCastiObce'):
                    sidlo_parts.append(city_part)
                if city := address_data.get('nazevObce'):
                    sidlo_parts.append(city)
                if psc := address_data.get('psc'):
                    sidlo_parts.insert(-1, f"{str(psc).zfill(5)}")
                
                sidlo_ares = ", ".join(filter(None, sidlo_parts))

            # Získání právní formy
            pravni_forma = None
            if pravni_forma_kod := basic_data.get('pravniForma'):
                pravni_forma = self._get_pravni_forma_name(pravni_forma_kod)

            company_data = AresCompanyData(
                ico=ico,
                nazev=basic_data.get('obchodniJmeno', ''),
                pravni_forma=pravni_forma,
                datum_vzniku=basic_data.get('datumVzniku', '').split('T')[0] if basic_data.get('datumVzniku') else None,
                sidlo=self._format_address(basic_data.get('sidlo', {})) if isinstance(basic_data.get('sidlo'), dict) else None,
                zamestnanci=employees,
                sidlo_ares=sidlo_ares,
                datumZaniku=basic_data.get('datumZaniku', '').split('T')[0] if basic_data.get('datumZaniku') else None
            )
            
            self.logger.debug(f"Získán počet zaměstnanců pro IČO {ico}: {employees}")
            self.logger.info(f"Úspěšně získána data pro IČO {ico}")
            return company_data

        except requests.exceptions.RequestException as e:
            self.logger.error(f"Síťová chyba při získávání dat z ARES pro IČO {ico}: {str(e)}")
        except Exception as e:
            self.logger.error(f"Neočekávaná chyba při získávání dat z ARES pro IČO {ico}: {str(e)}", exc_info=True)
        return None

    def _get_employee_category_name(self, kod: str) -> Optional[str]:
        """Převede kód kategorie počtu zaměstnanců na maximální počet"""
        KATEGORIE_POCTU_ZAMESTNANCU = {
            "000": "Neuvedeno",
            "110": "0",
            "120": "5",
            "130": "9",
            "210": "19",
            "220": "24",
            "230": "49",
            "240": "99",
            "310": "199",
            "320": "249",
            "330": "499",
            "340": "999",
            "410": "1499",
            "420": "1999",
            "430": "2499",
            "440": "2999",
            "450": "3999",
            "460": "4999",
            "470": "10000",
            "510": "10000+"
        }
        
        mapped_value = KATEGORIE_POCTU_ZAMESTNANCU.get(str(kod))
        if mapped_value:
            self.logger.debug(f"Úspěšně namapována kategorie zaměstnanců {kod} na {mapped_value}")
            return mapped_value
        else:
            self.logger.warning(f"Nenalezeno mapování pro kód kategorie zaměstnanců: {kod}")
            return None
    
    def _get_pravni_forma_name(self, kod: str) -> Optional[str]:
        """Převede kód právní formy na název"""
        # Mapování kódů na názvy
        PRAVNI_FORMY = {
            "101": "podnik FO",
            "102": "fyz",
            "111": "vos",
            "112": "sro",
            "113": "ks",
            "117": "nadace",
            "118": "nadační fond",
            "121": "as",
            "141": "obecně prospěšná společnost",
            "145": "společenství vlastníků jednotek",
            "151": "příspěvková organizace",
            "161": "ústav",
            "205": "družstvo",
            "301": "státní podnik",
            "325": "organizační složka státu",
            "331": "příspěvková organizace",
            "401": "živnostník",
            "421": "zahraniční fyzická osoba",
            "422": "zahraniční právnická osoba",
            "501": "odštěpný závod",
            "601": "vysoká škola",
            "641": "školská právnická osoba",
            "701": "sdružení",
            "704": "spolek",
            "706": "pobočný spolek",
            "736": "církevní organizace",
            "751": "zájmové sdružení",
            "801": "obec",
            "804": "kraj",
            "805": "hlavní město Praha",
        }
        return PRAVNI_FORMY.get(kod)

    def _format_address(self, address_data: dict) -> str:
        """Formátování adresy do standardního formátu"""
        if not address_data:
            return ''
        
        if address_data.get('textovaAdresa'):
            return address_data['textovaAdresa']
        
        parts = []
        if street := address_data.get('nazevUlice'):
            parts.append(street)
            
        if house_num := address_data.get('cisloDomovni'):
            num_str = str(house_num)
            if orientation_num := address_data.get('cisloOrientacni'):
                num_str += f"/{orientation_num}"
            parts.append(num_str)
            
        if city := address_data.get('nazevObce'):
            parts.append(city)
            
        if postal_code := address_data.get('psc'):
            parts.append(str(postal_code).zfill(5))
            
        return ', '.join(parts)

    def _format_ares_address(self, address_data: dict) -> str:
        """Formátování adresy pro ARES formát"""
        parts = []
        
        # Ulice a číslo
        if street := address_data.get('nazevUlice'):
            address_part = street
            if num := address_data.get('cisloDomovni'):
                address_part += f" {num}"
                if orient := address_data.get('cisloOrientacni'):
                    address_part += f"/{orient}"
            parts.append(address_part)
        elif num := address_data.get('cisloDomovni'):
            parts.append(f"č.p. {num}")
        
        # Část obce, pokud existuje
        if city_part := address_data.get('nazevCastiObce'):
            parts.append(city_part)
        
        # PSČ a město
        if psc := address_data.get('psc'):
            parts.append(f"{str(psc).zfill(5)}")
        if city := address_data.get('nazevObce'):
            parts.append(city)
            
        return ', '.join(parts)

@dataclass
class InsolvencniRizeni:
    """Třída reprezentující insolvenci subjektu"""
    spisovaZnacka: str
    soud: str
    datumZahajeni: datetime
    stavRizeni: str
    urlDetailRizeni: Optional[str] = None

class InsolvencniAnalyzer:
    """Třída pro analýzu insolvencí"""
    
    def __init__(self, logger=None):
        self.logger = logger or logging.getLogger(__name__)
        self._setup_logging()
        
    def _setup_logging(self):
        if not self.logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
            handler.setFormatter(formatter)
            self.logger.addHandler(handler)
            self.logger.setLevel(logging.INFO)

    def get_insolvence(self, ico: str) -> Optional[List[InsolvencniRizeni]]:
        """Získá informace o insolvencích subjektu"""
        try:
            url = f"https://ares.gov.cz/ekonomicke-subjekty-v-be/rest/ekonomicke-subjekty-vr/{ico}"
            response = requests.get(url)
            response.raise_for_status()
            data = response.json()
            
            insolvence = []
            
            for zaznam in data.get('zaznamy', []):
                for insolvencni_rizeni in zaznam.get('insolvence', []):
                    for zapis in insolvencni_rizeni.get('insolvencniZapis', []):
                        insolvence.append(InsolvencniRizeni(
                            spisovaZnacka=zapis.get('text', ''),
                            soud=data.get('soud', ''),
                            datumZahajeni=datetime.fromisoformat(zapis.get('datumZapisu').split('T')[0]),
                            stavRizeni=zapis.get('typZapisu', ''),
                            urlDetailRizeni=f"https://isir.justice.cz/isir/ueu/evidence_upadcu_detail.do?id={ico}"
                        ))

            return insolvence if insolvence else None

        except Exception as e:
            self.logger.error(f"Chyba při získávání informací o insolvenci pro IČO {ico}: {str(e)}")
            return None
            
class ZivefirmyScraper:
    def __init__(self, max_workers=2, contacted_companies_file=None):
        self.session = requests.Session()
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'cs,en-US;q=0.7,en;q=0.3',
            'Connection': 'keep-alive',
        }
        self._setup_logger()
        self.base_url = "https://www.zivefirmy.cz"
        self.max_workers = max_workers
        self.web_analyzer = WebAnalyzer(self.logger)
        self.ares_service = AresService(self.logger)
        
        # Inicializace filtru
        self.contacted_filter = None
        if contacted_companies_file:
            self.logger.info(f"Inicializuji filtr kontaktovaných firem ze souboru: {contacted_companies_file}")
            self.contacted_filter = ContactedCompaniesFilter(contacted_companies_file)
    
    def _setup_logger(self):
        self.logger = logging.getLogger('ZivefirmyScraper')
        self.logger.setLevel(logging.DEBUG)
        if not self.logger.handlers:
            fh = logging.FileHandler('zivefirmy_scraper.log', encoding='utf-8')
            fh.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
            self.logger.addHandler(fh)

    def _format_phone_number(self, phone: str) -> str:
        """Formátuje telefonní číslo do standardního formátu"""
        # Odstranění všech nečíselných znaků kromě +
        cleaned = re.sub(r'[^\d+]', '', phone)
        
        # Pro česká čísla bez předvolby přidáme +420
        if len(cleaned) == 9 and cleaned.isdigit():
            cleaned = '+420' + cleaned
        
        # Formátování čísla s předvolbou
        if len(cleaned) >= 12 and cleaned.startswith('+'):
            prefix = cleaned[:4]  # +420
            rest = cleaned[4:]
            if len(rest) == 9:
                return f"{prefix} {rest[:3]} {rest[3:6]} {rest[6:]}"
        
        # Pokud formátování selže, vrátíme původní číslo
        return phone

    def get_soup(self, url: str) -> BeautifulSoup:
        """Vylepšená metoda pro získání HTML obsahu s exponenciálním backoff"""
        max_retries = 3
        base_delay = 2
        
        for attempt in range(max_retries):
            try:
                # Přidáme náhodné zpoždění mezi requesty
                time.sleep(random.uniform(2, 3))
                
                response = self.session.get(url, headers=self.headers, timeout=30)
                response.raise_for_status()
                return BeautifulSoup(response.text, 'html.parser')
                
            except Exception as e:
                self.logger.error(f"Pokus {attempt + 1}/{max_retries} selhal pro {url}: {str(e)}")
                if attempt < max_retries - 1:
                    # Exponenciální backoff
                    delay = base_delay * (2 ** attempt) + random.uniform(0, 2)
                    time.sleep(delay)
                    
        return None
    
    def find_company_website(self, company_name: str, address: str) -> str:
        try:
            # Vyčistíme název firmy a získáme město
            company_name = company_name.replace(',', '').replace('.', '')
            city = address.split()[-1] if address else ''
            
            search_query = f"{company_name} {city}"
            encoded_query = urllib.parse.quote(search_query)
            url = f"https://www.google.com/search?q={encoded_query}"
            
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True)
                page = browser.new_page()
                page.goto(url, timeout=30000)
                
                results = page.query_selector_all('div.g')
                
                for result in results[:3]:
                    link = result.query_selector('a')
                    if link:
                        href = link.get_attribute('href')
                        if href and not any(x in href for x in ['facebook.com', 'instagram.com', 'linkedin.com', 'twitter.com', 'zivefirmy.cz', 'firmy.cz']):
                            browser.close()
                            return href
                            
                browser.close()
                return None
                
        except Exception as e:
            self.logger.error(f"Chyba při hledání webu pro {company_name}: {str(e)}")
            return None

    def get_total_pages(self, soup: BeautifulSoup) -> int:
        try:
            pagination = soup.find('ul', class_='pagination')
            if pagination:
                last_page = pagination.find_all('a')[-2].text
                return int(last_page)
        except Exception as e:
            self.logger.error(f"Chyba při získávání počtu stránek: {e}")
        return 1

    def get_company_details(self, url: str) -> dict:
        """Získá kompletní detaily o firmě včetně PSČ"""
        max_retries = 3
        base_delay = 3
        company_data = {}
        
        for attempt in range(max_retries):
            try:
                self.logger.info(f"Začínám získávání detailů firmy z URL: {url}")
                soup = self.get_soup(url)
                if not soup:
                    raise Exception("Nepodařilo se získat HTML obsah")

                company_data = {
                    'nazev': self._get_company_name(soup),
                    'ic': self._get_ico(soup),
                    'adresa': self._get_address(soup),
                    'emaily': self._get_emails(soup),
                    'web': self._get_websites(soup), 
                    'popis': self._get_description(soup),
                    'kontaktni_osoby': self._get_contact_persons(soup),
                    'provozni_doba': self._get_opening_hours(soup),
                    'kategorie_cinnosti': self._get_company_activities(soup),
                    'cinnosti': self._get_company_activities(soup),  
                    'url': url
                }

                # Pokus získat PSČ přímo ze stránky
                psc = self._get_psc(soup)
                if psc:
                    company_data['psc'] = psc
                    self.logger.info(f"PSČ získáno ze stránky: {psc}")

                # Pokud máme web, analyzujeme ho
                if company_data.get('web'):
                    web_urls = company_data['web']
                    if isinstance(web_urls, str):
                        web_urls = [url.strip() for url in web_urls.split() if url.strip()]
                    
                    self.logger.info(f"Nalezeny weby pro firmu: {web_urls}")
                    
                    for web_url in web_urls:
                        try:
                            web_analysis = self.web_analyzer.analyze_website(web_url)
                            if web_analysis:
                                company_data.update(web_analysis)
                                break
                        except Exception as e:
                            self.logger.error(f"Chyba při analýze webu {web_url}: {str(e)}")
                            continue

                if not company_data.get('popis'):
                    popis = []
                    for activity, points in company_data['cinnosti'].items():
                        popis.append(f"**{activity}**")
                        if points:
                            popis.extend([f"* {point}" for point in points])
                    company_data['popis'] = "\n".join(popis) if popis else ""
                # Pokud nemáme web, pokusíme se ho najít přes Google
                
                if not company_data.get('web'):
                    city = re.search(r'([^,]+)(?:,[^,]+)?$', company_data.get('adresa', '')).group(1).strip() if company_data.get('adresa') else ''
                    found_website = self.find_company_website(company_data['nazev'], city)
                    if found_website:
                        company_data['web'] = [found_website]
                        try:
                            web_analysis = self.web_analyzer.analyze_website(found_website)
                            if web_analysis:
                                company_data.update(web_analysis)
                        except Exception as e:
                            self.logger.error(f"Chyba při analýze nalezeného webu {found_website}: {str(e)}")

                # Pokud máme IČO, získáme data z ARESu
                if company_data.get('ic'):
                    ares_data = self.ares_service.get_company_details(company_data['ic'])
                    if ares_data:
                        # Pokud jsme nezískali PSČ ze stránky, zkusíme ho získat z ARESu
                        if not company_data.get('psc') and ares_data.sidlo_ares:
                            psc_match = re.search(r'\b\d{3}\s*\d{2}\b', ares_data.sidlo_ares)
                            if psc_match:
                                company_data['psc'] = psc_match.group().replace(' ', '')

                        # Aktualizace dalších dat z ARESu
                        company_data.update({
                            'zamestnanci': ares_data.zamestnanci,
                            'datum_vzniku': ares_data.datum_vzniku,
                            'pravni_forma': ares_data.pravni_forma,
                            'sidlo_ares': ares_data.sidlo
                        })

                return self._clean_company_data(company_data)

            except Exception as e:
                error_msg = f"Pokus {attempt + 1}/{max_retries} selhal pro {url}: {str(e)}"
                if attempt == max_retries - 1:
                    self.logger.error(error_msg)
                else:
                    self.logger.warning(error_msg)
                    time.sleep(base_delay * (attempt + 1) + random.uniform(0, 2))

        return company_data

    def _get_psc(self, soup: BeautifulSoup) -> Optional[str]:
        """Získá PSČ z HTML stránky"""
        try:
            # Hledání v meta datech
            meta_desc = soup.find('meta', {'name': 'description'})
            if meta_desc:
                desc_content = meta_desc.get('content', '')
                psc_match = re.search(r'\b\d{3}\s*\d{2}\b', desc_content)
                if psc_match:
                    return psc_match.group().replace(' ', '')
            
            # Hledání v adrese
            address_wrapper = soup.find('div', class_='wrapper-left')
            if address_wrapper:
                text = address_wrapper.get_text()
                psc_match = re.search(r'\b\d{3}\s*\d{2}\b', text)
                if psc_match:
                    return psc_match.group().replace(' ', '')

            # Hledání v celém textu stránky jako záložní řešení
            text = soup.get_text()
            psc_matches = re.findall(r'\b\d{3}\s*\d{2}\b', text)
            if psc_matches:
                # Vracíme první nalezené PSČ
                return psc_matches[0].replace(' ', '')
                    
            return None
        except Exception as e:
            self.logger.error(f"Chyba při získávání PSČ: {e}")
            return None
        
    def _get_basic_data(self, soup: BeautifulSoup, url: str) -> dict:
        """Získá základní data o firmě z HTML"""
        return {
            'nazev': self._get_company_name(soup),
            'ic': self._get_ico(soup),
            'adresa': self._get_address(soup),
            'emaily': self._get_emails(soup),
            'web': self._get_websites(soup), 
            'popis': self._get_description(soup),
            'kontaktni_osoby': self._get_contact_persons(soup),
            'provozni_doba': self._get_opening_hours(soup),
            'kategorie_cinnosti': self._get_company_activities(soup),  # Změněný klíč
            'url': url
        }

    def _validate_required_fields(self, data: dict):
        """Ověří přítomnost povinných polí"""
        required_fields = ['nazev', 'ic', 'adresa']
        missing_fields = [field for field in required_fields if not data.get(field)]
        
        if missing_fields:
            raise ValueError(f"Chybí povinná pole: {', '.join(missing_fields)}")

    

    def _analyze_company_website(self, data: dict) -> dict:
        """Provede analýzu webových stránek firmy"""
        if not data.get('web'):
            return data
            
        for web_url in data['web']:
            try:
                web_analysis = self.web_analyzer.analyze_website(web_url)
                if any(web_analysis.values()):
                    data.update({
                        'technologie': web_analysis['technologies'],
                        'certifikace': web_analysis['certifications'],
                        'jazyky': web_analysis['languages']
                    })
                    break
            except Exception as e:
                self.logger.error(f"Chyba při analýze webu {web_url}: {e}")
        
        return data

    def _clean_company_data(self, data: dict) -> dict:
        """Vyčistí data od prázdných hodnot a provede případné formátování"""
        # Odstranění None hodnot a prázdných seznamů/řetězců
        cleaned_data = {k: v for k, v in data.items() 
                    if v is not None and v != "" and v != [] and v != {}}
        
        # Případné další čištění nebo formátování dat
        if 'adresa' in cleaned_data:
            cleaned_data['adresa'] = cleaned_data['adresa'].strip()
            
        # Přidejte tuto část pro čištění jazykových dat
        if 'language_versions' in cleaned_data and not cleaned_data['language_versions']:
            del cleaned_data['language_versions']
        if 'languages' in cleaned_data and not cleaned_data['languages']:
            del cleaned_data['languages']
        
        return cleaned_data

    def _get_company_name(self, soup: BeautifulSoup) -> str:
        try:
            return soup.find('h1').text.strip()
        except:
            return ""

    def _get_ico(self, soup: BeautifulSoup) -> str:
        try:
            ico_elem = soup.find('span', class_='text-ico', string='IČ')
            if ico_elem:
                return ico_elem.find_next('span').text.strip()
        except:
            return ""
        return ""


    def _get_address(self, soup: BeautifulSoup) -> str:
        """Vylepšená metoda pro získávání adresy"""
        try:
            # Primární způsob - z wrapper-left
            wrapper_left = soup.find('div', class_='wrapper-left')
            if wrapper_left:
                # Získáme všechny textové části včetně odkazů
                address_parts = []
                for element in wrapper_left.children:
                    if element.name == 'a':
                        address_parts.append(element.text.strip())
                    elif isinstance(element, str) and element.strip():
                        address_parts.append(element.strip())
                    elif element.name == 'br':
                        continue
                
                # Vyčistíme a spojíme části adresy
                address = ' '.join([part for part in address_parts if part and part != ','])
                return address.replace('\n', ' ').strip()
                
            # Záložní způsob - z meta description
            meta_desc = soup.find('meta', {'name': 'description'})
            if meta_desc:
                desc_content = meta_desc.get('content', '')
                match = re.search(r'(?:s\.r\.o\.|a\.s\.|v\.o\.s\.|k\.s\.|spol\.\s+s\s+r\.o\.|družstvo),\s*([^,]*,[^,]*(?:,[^,]*)?)', desc_content)
                if match:
                    return match.group(1).strip()
                    
            # Další záložní způsob - zkusit najít adresu v textu stránky
            address_section = soup.find(string=re.compile(r'(Adresa|Sídlo)'))
            if address_section:
                parent = address_section.find_parent('div')
                if parent:
                    address_text = parent.get_text(strip=True)
                    # Vyčistit text od názvů sekcí a dalších nepotřebných částí
                    address_text = re.sub(r'^(Adresa|Sídlo):', '', address_text)
                    return address_text.strip()
                    
        except Exception as e:
            self.logger.error(f"Chyba při získávání adresy: {e}")
        return ""

    def _get_emails(self, soup: BeautifulSoup) -> list:
        emails = []
        blocked_emails = ['zivefirmy@databox.cz']
        try:
            email_links = soup.find_all('a', href=re.compile(r'^mailto:'))
            for link in email_links:
                email = link['href'].replace('mailto:', '')
                if email and email not in blocked_emails:
                    emails.append(email)
        except Exception as e:
            self.logger.error(f"Chyba při získávání emailů: {e}")
        return emails
    
    def _get_websites(self, soup: BeautifulSoup) -> list:
        """Získá webové stránky firmy jako list"""
        websites = []
        blocked_domains = [
            'facebook.com',
            'instagram.com',
            'linkedin.com',
            'twitter.com',
            'firmy.cz',
            'mapy.cz'
        ]
        
        try:
            web_links = soup.find_all('a', class_='link-web')
            for link in web_links:
                href = link.get('href')
                if href and not href.startswith('mailto:'):
                    websites.append(href.strip())
                
            self.logger.debug(f"Nalezené weby: {websites}")
            
        except Exception as e:
            self.logger.error(f"Chyba při získávání webů: {str(e)}")
        return websites


    def _get_company_activities(self, soup: BeautifulSoup) -> dict:
        """Extracts company activities along with their descriptions"""
        activities = {}
        activities_section = soup.find('div', class_='cinnosti')
        if activities_section:
            for item in activities_section.find_all('div', class_='item'):
                title = item.find('div', class_='title')
                if title:
                    activity_name = title.text.strip()
                    activities[activity_name] = []
                    
                    # Extract bullet points
                    points = item.find_all('li')
                    for point in points:
                        activities[activity_name].append(point.text.strip())
                    
        return activities
    
    def _get_description(self, soup: BeautifulSoup) -> str:
        try:
            desc_elem = soup.find('p', class_='desc')
            if desc_elem:
                return desc_elem.text.strip()
        except:
            return ""
        return ""

    def _get_contact_persons(self, soup: BeautifulSoup) -> list:
        """Získá a seřadí kontaktní osoby a telefonní čísla - kontakty s popisky budou první"""
        # Používáme slovník pro sledování jedinečných kontaktů
        contacts_dict = {}
        
        try:
            # 1. Zpracování kontaktních osob
            contact_section = soup.find('div', class_='title-section', string='Kontaktní osoby')
            if contact_section:
                contact_part = contact_section.find_next('div', class_='part')
                if contact_part:
                    for person in contact_part.find_all('li'):
                        person_info = person.text.strip()
                        if person_info and not any(person_info in v for v in contacts_dict.values()):
                            # Použijeme prioritu 1 pro kontakty s popisem
                            contacts_dict[('1', 'person_' + str(len(contacts_dict)))] = person_info

            # 2. Zpracování telefonních kontaktů
            contacts_section = soup.find('div', class_='title-section', string='Kontakty')
            if contacts_section:
                contacts_part = contacts_section.find_next('div', class_='part')
                if contacts_part:
                    # Zpracování elementů action-copy
                    phone_elements = contacts_part.find_all('span', class_='action-copy')
                    for element in phone_elements:
                        phone = element.get('data-text', '').strip()
                        if phone:
                            # Vyčistíme a zformátujeme telefon
                            cleaned_phone = self._format_phone_number(phone)
                            note = element.find_next('span', class_='text-sm')
                            if note:
                                note_text = note.text.strip()
                                # Priorita 1 pro kontakty s popisem
                                contacts_dict[('1', cleaned_phone)] = f"{cleaned_phone} - {note_text}"
                            else:
                                # Priorita 2 pro kontakty bez popisu
                                contacts_dict[('2', cleaned_phone)] = cleaned_phone

            # 3. Záložní prohledání celé stránky pro elementy action-copy
            all_phone_elements = soup.find_all('span', class_='action-copy')
            for element in all_phone_elements:
                if 'data-text' in element.attrs:
                    phone = element['data-text'].strip()
                    if re.match(r'^\+?\d[\d\s-]+$', phone):
                        cleaned_phone = self._format_phone_number(phone)
                        if not any(cleaned_phone in v for v in contacts_dict.values()):
                            note = element.find_next('span', class_='text-sm')
                            if note:
                                note_text = note.text.strip()
                                # Priorita 1 pro kontakty s popisem
                                contacts_dict[('1', cleaned_phone)] = f"{cleaned_phone} - {note_text}"
                            else:
                                # Priorita 2 pro kontakty bez popisu
                                contacts_dict[('2', cleaned_phone)] = cleaned_phone

        except Exception as e:
            self.logger.error(f"Chyba při získávání kontaktních osob: {e}")

        # Seřadíme kontakty podle priority (1 = s popisem, 2 = bez popisu) a vrátíme pouze hodnoty
        return [v for k, v in sorted(contacts_dict.items())]

    def _get_opening_hours(self, soup: BeautifulSoup) -> str:
        """Získá a zformátuje provozní dobu včetně polední pauzy"""
        try:
            hours_section = None
            for section in soup.find_all('div', class_='title-section'):
                if 'Provozní doba' in section.text:
                    hours_section = section
                    break
                    
            if hours_section:
                part_div = hours_section.find_next('div', class_='part')
                if part_div:
                    # Získáme text a rozdělíme na řádky
                    hours_text = part_div.get_text(strip=True)
                    if not hours_text:
                        return "Neuvedeno"
                    
                    # Vytvoříme slovník pro ukládání časů podle dnů
                    schedule = {}
                    lunch_breaks = {}
                    
                    # Regulární výraz pro nalezení dnů a časů včetně polední pauzy
                    pattern = r'(PO|ÚT|ST|ČT|PÁ|SO|NE)[:\s]+(\d{1,2}[:\.]\d{2})\s*-\s*(\d{1,2}[:\.]\d{2})(?:\s*,?\s*(\d{1,2}[:\.]\d{2})\s*-\s*(\d{1,2}[:\.]\d{2}))?\s*(?:\((?:polední\s+)?pauza\s+(\d{1,2}[:\.]\d{2})\s*-\s*(\d{1,2}[:\.]\d{2})\))?'
                    matches = re.finditer(pattern, hours_text, re.IGNORECASE)
                    
                    for match in matches:
                        groups = match.groups()
                        day = groups[0]
                        
                        # Standardizace formátu času (tečka na dvojtečku)
                        times = []
                        if groups[1] and groups[2]:  # První časový úsek
                            start1 = groups[1].replace('.', ':')
                            end1 = groups[2].replace('.', ':')
                            times.append(f"{start1}-{end1}")
                        
                        if groups[3] and groups[4]:  # Druhý časový úsek
                            start2 = groups[3].replace('.', ':')
                            end2 = groups[4].replace('.', ':')
                            times.append(f"{start2}-{end2}")
                        
                        schedule[day] = ", ".join(times)
                        
                        # Přidání polední pauzy
                        if groups[5] and groups[6]:  # Polední pauza
                            lunch_start = groups[5].replace('.', ':')
                            lunch_end = groups[6].replace('.', ':')
                            lunch_breaks[day] = f"{lunch_start}-{lunch_end}"
                    
                    # Seskupíme dny se stejnou provozní dobou a pauzou
                    grouped_schedule = {}
                    for day, time_range in schedule.items():
                        lunch_break = lunch_breaks.get(day, '')
                        key = (time_range, lunch_break)
                        if key in grouped_schedule:
                            grouped_schedule[key].append(day)
                        else:
                            grouped_schedule[key] = [day]
                    
                    # Formátování výstupu
                    formatted_hours = []
                    for (time_range, lunch_break), days in grouped_schedule.items():
                        # Seřadíme dny podle pořadí v týdnu
                        day_order = {'PO': 1, 'ÚT': 2, 'ST': 3, 'ČT': 4, 'PÁ': 5, 'SO': 6, 'NE': 7}
                        days.sort(key=lambda x: day_order[x])
                        
                        if len(days) > 2 and day_order[days[-1]] - day_order[days[0]] == len(days) - 1:
                            days_str = f"{days[0]}-{days[-1]}"
                        else:
                            days_str = ", ".join(days)
                        
                        hour_str = f"{days_str}: {time_range}"
                        if lunch_break:
                            hour_str += f" (polední pauza {lunch_break})"
                        
                        formatted_hours.append(hour_str)
                    
                    return " | ".join(formatted_hours)
                
            return "Neuvedeno"
            
        except Exception as e:
            self.logger.error(f"Chyba při získávání provozní doby: {e}")
            return "Neuvedeno"
    
    def scrape_category(self, category_url: str, max_companies: int = None) -> list:
        companies = []
        failed_urls = []
        
        try:
            initial_soup = self.get_soup(category_url)
            if not initial_soup:
                return companies

            total_pages = self.get_total_pages(initial_soup)
            total_companies_estimate = min(max_companies, total_pages * 40) if max_companies else total_pages * 40
            
            with tqdm(total=total_companies_estimate, desc="Stahování firem", unit="firma") as pbar:
                for page in range(1, total_pages + 1):
                    if max_companies and len(companies) >= max_companies:
                        break
                        
                    page_url = f"{category_url}&pg={page}" if page > 1 else category_url
                    time.sleep(random.uniform(2, 3))
                    
                    soup = self.get_soup(page_url)
                    if not soup:
                        continue

                    company_items = soup.find_all('div', class_='company-item')
                    page_urls = []
                    
                    for item in company_items:
                        if max_companies and len(companies) + len(page_urls) >= max_companies:
                            break
                            
                        title_link = item.find('div', class_='title').find('a')
                        if title_link and 'href' in title_link.attrs:
                            company_url = urljoin(self.base_url, title_link['href'])
                            page_urls.append(company_url)

                    with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                        future_to_url = {executor.submit(self.get_company_details, url): url 
                                    for url in page_urls}
                        
                        for future in as_completed(future_to_url):
                            url = future_to_url[future]
                            try:
                                company_data = future.result()
                                if company_data and company_data.get('ic'):
                                    # Get ARES data for each company
                                    ares_data = self.ares_service.get_company_details(company_data['ic'])
                                    if ares_data:
                                        company_data.update({
                                            'zamestnanci': ares_data.zamestnanci,
                                            'datum_vzniku': ares_data.datum_vzniku,
                                            'pravni_forma': ares_data.pravni_forma,
                                            'sidlo_ares': ares_data.sidlo
                                        })
                                    companies.append(company_data)
                                    pbar.update(1)
                                    pbar.set_description(f"Staženo {len(companies)} firem")
                                else:
                                    failed_urls.append(url)
                            except Exception as e:
                                self.logger.error(f"Chyba při zpracování {url}: {str(e)}")
                                failed_urls.append(url)

                if failed_urls:
                    self.logger.info(f"Opakuji stažení {len(failed_urls)} neúspěšných URL")
                    for url in failed_urls:
                        try:
                            time.sleep(random.uniform(2, 5))
                            company_data = self.get_company_details(url)
                            if company_data:
                                companies.append(company_data)
                        except Exception as e:
                            self.logger.error(f"Neúspěšný druhý pokus pro {url}: {str(e)}")

                self.logger.info(f"Celkem staženo {len(companies)} firem")
                return companies

        except Exception as e:
            self.logger.error(f"Chyba při scrapování kategorie: {str(e)}")
            return companies

def format_phone_number(phone: str) -> str:
    """Formátuje pouze 9místná a 12místná telefonní čísla"""
    cleaned = re.sub(r'[^\d+]', '', phone)
    
    if len(cleaned) == 13 and cleaned.startswith('+'):
        prefix = cleaned[:4]
        rest = cleaned[4:]
        if len(rest) == 9:
            return f"{prefix} {rest[:3]} {rest[3:6]} {rest[6:]}"
    elif len(cleaned) == 9 and cleaned.isdigit():
        return f"{cleaned[:3]} {cleaned[3:6]} {cleaned[6:]}"
    return phone

def main():
    while True:
        try:
            max_workers = int(input("Zadejte počet paralelních vláken (doporučeno 3-5): "))
            if 1 <= max_workers <= 10:
                break
            print("Prosím zadejte číslo mezi 1 a 10")
        except ValueError:
            print("Prosím zadejte platné číslo")

    # Inicializace filtru
    contacted_file = input("Zadejte cestu k souboru s již kontaktovanými firmami (Enter pro přeskočení): ").strip()
    
    if contacted_file and os.path.exists(contacted_file):
        print(f"\nInicializuji filtr kontaktovaných firem ze souboru: {contacted_file}")
        try:
            scraper = ZivefirmyScraper(
                max_workers=max_workers,
                contacted_companies_file=contacted_file
            )
            if scraper.contacted_filter:
                print(f"Úspěšně načteno {len(scraper.contacted_filter.contacted_companies)} kontaktovaných firem")
                print("Seznam načtených firem pro kontrolu:")
                for company in list(scraper.contacted_filter.contacted_companies)[:5]:
                    print(f"- {company}")
                if len(scraper.contacted_filter.contacted_companies) > 5:
                    print(f"... a dalších {len(scraper.contacted_filter.contacted_companies) - 5} firem")
            else:
                print("VAROVÁNÍ: Filtr se neinicializoval správně!")
                return
        except Exception as e:
            print(f"Chyba při inicializaci filtru: {str(e)}")
            return
    else:
        scraper = ZivefirmyScraper(max_workers=max_workers)
        if contacted_file:
            print(f"VAROVÁNÍ: Soubor {contacted_file} neexistuje!")
            return

    category_url = "https://www.zivefirmy.cz/strojirenstvi-kovo_t1144?loc=1"
    
    print("\nNačítám informace o kategorii...")
    initial_soup = scraper.get_soup(category_url)
    if not initial_soup:
        print("Nepodařilo se načíst stránku")
        return

    total_pages = scraper.get_total_pages(initial_soup)
    total_companies = total_pages * 40  # Předpokládáme 40 firem na stránku
    print(f"Nalezeno přibližně {total_companies} firem v této kategorii")
    
    # Nová část - dotaz na způsob stahování
    while True:
        start_mode = input("Chcete začít od začátku (Z) nebo pokračovat od určité pozice (P)? : ")
        if start_mode in ['Z', 'P']:
            break
        print("Prosím zadejte Z nebo P")
    
    start_position = 0
    if start_mode == 'P':
        while True:
            try:
                start_position = int(input(f"Od které pozice chcete pokračovat (1-{total_companies}): ")) - 1
                if 0 <= start_position < total_companies:
                    break
                print(f"Prosím zadejte číslo mezi 1 a {total_companies}")
            except ValueError:
                print("Prosím zadejte platné číslo")
    
    while True:
        response = input("Stáhnout všechny zbývající firmy z této kategorie? (y/n): ").lower()
        if response == 'y':
            max_companies = total_companies - start_position
            break
        elif response == 'n':
            while True:
                remaining_companies = total_companies - start_position
                try:
                    max_companies = int(input(f"Kolik firem chcete stáhnout (max {remaining_companies}): "))
                    if 0 < max_companies <= remaining_companies:
                        break
                    print(f"Prosím zadejte číslo mezi 1 a {remaining_companies}")
                except ValueError:
                    print("Prosím zadejte platné číslo")
            break
        else:
            print("Prosím odpovězte 'y' nebo 'n'")

    print(f"\nZačínám stahování {max_companies} firem od pozice {start_position + 1}...")
    
    companies = []
    with tqdm(total=max_companies, desc="Stahování firem", unit="firma") as pbar:
        for page in range(start_position // 40 + 1, total_pages + 1):
            if len(companies) >= max_companies:
                break
                
            page_url = f"{category_url}&pg={page}" if page > 1 else category_url
            time.sleep(random.uniform(2, 4))
            
            soup = scraper.get_soup(page_url)
            if not soup:
                continue

            company_items = soup.find_all('div', class_='company-item')
            
            # Aplikace offsetu pouze na první stránce
            if page == start_position // 40 + 1:
                offset = start_position % 40
                company_items = company_items[offset:]
            
            page_urls = []
            for item in company_items:
                if len(companies) + len(page_urls) >= max_companies:
                    break
                    
                title_link = item.find('div', class_='title').find('a')
                if title_link and 'href' in title_link.attrs:
                    company_url = urljoin(scraper.base_url, title_link['href'])
                    page_urls.append(company_url)

            with ThreadPoolExecutor(max_workers=scraper.max_workers) as executor:
                future_to_url = {executor.submit(scraper.get_company_details, url): url 
                            for url in page_urls}
                
                for future in as_completed(future_to_url):
                    url = future_to_url[future]
                    try:
                        company_data = future.result()
                        if company_data and company_data.get('ic'):
                            # Get ARES data for each company
                            ares_data = scraper.ares_service.get_company_details(company_data['ic'])
                            if ares_data:
                                company_data.update({
                                    'zamestnanci': ares_data.zamestnanci,
                                    'datum_vzniku': ares_data.datum_vzniku,
                                    'pravni_forma': ares_data.pravni_forma,
                                    'sidlo_ares': ares_data.sidlo
                                })
                            companies.append(company_data)
                            pbar.update(1)
                            pbar.set_description(f"Staženo {len(companies)} firem")
                    except Exception as e:
                        print(f"Chyba při zpracování firmy {url}: {str(e)}")
                        continue

    if companies:
        print(f"\nStaženo {len(companies)} firem")
        
        # Export dat
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        start_position_str = f"_start{start_position + 1}" if start_position > 0 else ""
        
        while True:
            save_format = input("Uložit jako Excel? (y/n): ").lower()
            if save_format == 'y':
                try:
                    excel_file = f"strojirenske_firmy_{len(companies)}{start_position_str}_{timestamp}.xlsx"
                    
                    # Inicializace DataExporteru a export dat
                    data_exporter = DataExporter(scraper.logger)
                    data_exporter.export_to_excel(companies, excel_file)
                    
                    print(f"Data úspěšně exportována do {excel_file}")
                    print(f"Soubor obsahuje detailní informace o firmách včetně exportních aktivit a souhrnné statistiky")
                    break
                    
                except Exception as e:
                    print(f"Chyba při ukládání dat: {e}")
                    continue
            elif save_format == 'n':
                try:
                    json_file = f"strojirenske_firmy_{len(companies)}{start_position_str}_{timestamp}.json"
                    with open(json_file, 'w', encoding='utf-8') as f:
                        json.dump(companies, f, ensure_ascii=False, indent=2)
                    print(f"Data uložena do {json_file}")
                    break
                except Exception as e:
                    print(f"Chyba při ukládání do JSONu: {e}")
                    continue
            else:
                print("Prosím odpovězte 'y' nebo 'n'")
    else:
        print("Nebyly nalezeny žádné firmy!")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nStahování přerušeno uživatelem")
    except Exception as e:
        print(f"\nChyba: {e}")
        logging.error(f"Neočekávaná chyba: {str(e)}", exc_info=True)